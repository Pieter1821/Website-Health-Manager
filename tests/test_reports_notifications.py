"""Report + notification helper tests (no network)."""

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from whm.domain.models import (
    Finding,
    FindingStatus,
    HealthCheckResult,
    HealthStatus,
    RiskLevel,
    Website,
)
from whm.infrastructure.notifications import format_message, should_notify
from whm.infrastructure.reports import (
    build_csv_report,
    build_excel_report,
    build_portfolio_excel_report,
    export_csv,
    export_excel,
    save_portfolio_report_to_downloads,
    save_report_to_downloads,
)


def _sample():
    site = Website(
        url="https://example.com",
        domain="example.com",
        display_name="Example",
        id=1,
    )
    result = HealthCheckResult(
        website_id=1,
        overall_status=HealthStatus.CRITICAL,
        risk_level=RiskLevel.HIGH,
        website_status=HealthStatus.HEALTHY,
        ssl_status=HealthStatus.HEALTHY,
        domain_status=HealthStatus.UNKNOWN,
        dns_status=HealthStatus.HEALTHY,
        email_status=HealthStatus.CRITICAL,
        findings=[
            Finding(
                "ssl",
                "Certificate expiring soon",
                FindingStatus.INCORRECT,
                "14 days",
                "Renew the certificate",
            ),
            Finding("spf", "SPF missing", FindingStatus.MISSING, "none", "Add SPF"),
            Finding(
                "security",
                "HSTS missing",
                FindingStatus.MISSING,
                "ignored",
                "should not appear",
            ),
        ],
        raw={
            "ssl": {"not_after": "2026-09-01T12:00:00+00:00", "days_remaining": 30},
            "whois": {"expiration_date": "2027-01-15", "days_remaining": 166},
        },
    )
    return site, result


def test_should_notify_levels():
    _, result = _sample()
    assert should_notify(result, "critical")
    assert should_notify(result, "warning")
    assert should_notify(result, "always")
    assert not should_notify(result, "never")


def test_format_message_plain():
    site, result = _sample()
    text = format_message(site, result)
    assert "Example" in text
    assert "needs a fix" in text.lower()


def test_export_excel_and_csv(tmp_path: Path):
    site, result = _sample()
    xlsx = export_excel(tmp_path / "r.xlsx", site, result)
    csv_path = export_csv(tmp_path / "r.csv", site, result)
    assert xlsx.exists()
    wb = load_workbook(xlsx)
    assert wb.sheetnames == ["Summary", "Problems to fix"]
    assert wb["Problems to fix"]["C2"].value == "Certificate expiring soon"
    assert "SPF missing" not in "".join(
        str(c.value or "") for row in wb["Problems to fix"].iter_rows(min_row=2, max_col=5) for c in row
    )
    text = csv_path.read_text(encoding="utf-8-sig")
    assert "Certificate expiring soon" in text
    assert "Renew the certificate" in text
    assert "SSL expires" in text
    assert "2026-09-01 (30 days left)" in text
    assert "Domain expires" in text
    assert "2027-01-15 (166 days left)" in text
    assert "SPF missing" not in text
    assert "HSTS missing" not in text
    assert "\nSummary,Email," not in text
    summary = wb["Summary"]
    assert any(
        c.value == "SSL expires" and summary.cell(row=c.row, column=2).value
        == "2026-09-01 (30 days left)"
        for row in summary.iter_rows(min_col=1, max_col=1)
        for c in row
    )
    assert any(
        c.value == "Domain expires" and summary.cell(row=c.row, column=2).value
        == "2027-01-15 (166 days left)"
        for row in summary.iter_rows(min_col=1, max_col=1)
        for c in row
    )


def test_build_excel_and_csv_bytes():
    site, result = _sample()
    xname, xbytes = build_excel_report(site, result)
    cname, cbytes = build_csv_report(site, result)
    assert xname.endswith(".xlsx")
    assert cname.endswith(".csv")
    assert xbytes[:2] == b"PK"
    assert b"Certificate expiring soon" in cbytes
    assert b"SPF missing" not in cbytes
    wb = load_workbook(BytesIO(xbytes))
    assert wb["Summary"]["B3"].value == "Example"


def test_dispatch_notifications_respects_never(monkeypatch):
    from whm.infrastructure.notifications import dispatch_notifications

    site, result = _sample()
    sent = dispatch_notifications(site, result, {"notify_on": "never", "notify_desktop": "1"})
    assert sent == []


def test_dispatch_notifications_channels(monkeypatch):
    from whm.infrastructure import notifications
    from whm.infrastructure.notifications import dispatch_notifications

    site, result = _sample()
    posts: list[tuple[str, dict]] = []

    monkeypatch.setattr(notifications, "notify_desktop", lambda *a, **k: None)
    monkeypatch.setattr(
        notifications,
        "_post_json",
        lambda url, payload, timeout=10.0: posts.append((url, payload)),
    )

    sent = dispatch_notifications(
        site,
        result,
        {
            "notify_on": "critical",
            "notify_desktop": "1",
            "slack_webhook": "https://hooks.example/slack",
            "discord_webhook": "https://hooks.example/discord",
            "teams_webhook": "https://hooks.example/teams",
            "generic_webhook": "https://hooks.example/generic",
        },
    )
    assert "desktop" in sent
    assert "slack" in sent
    assert "discord" in sent
    assert "teams" in sent
    assert "webhook" in sent
    assert any(p[1].get("text") for p in posts if "slack" in p[0])
    assert any(p[1].get("content") for p in posts if "discord" in p[0])
    assert any(p[1].get("@type") == "MessageCard" for p in posts)
    assert any(p[1].get("source") == "website-health-manager" for p in posts)


def test_dispatch_partial_failure_still_sends_others(monkeypatch):
    from whm.infrastructure import notifications
    from whm.infrastructure.notifications import dispatch_notifications

    site, result = _sample()

    def fail_slack(webhook, text):
        raise TimeoutError("nope")

    monkeypatch.setattr(notifications, "notify_desktop", lambda *a, **k: None)
    monkeypatch.setattr(notifications, "notify_slack", fail_slack)
    monkeypatch.setattr(notifications, "notify_discord", lambda *a, **k: None)

    sent = dispatch_notifications(
        site,
        result,
        {
            "notify_on": "critical",
            "notify_desktop": "1",
            "slack_webhook": "https://hooks.example/slack",
            "discord_webhook": "https://hooks.example/discord",
        },
    )
    assert "desktop" in sent
    assert "discord" in sent
    assert "slack" not in sent


def test_portfolio_excel_includes_all_sites(tmp_path: Path, monkeypatch):
    site, result = _sample()
    unchecked = Website(
        url="https://other.example",
        domain="other.example",
        display_name="Other",
        id=2,
    )
    name, payload = build_portfolio_excel_report([(site, result), (unchecked, None)])
    assert name.startswith("whm-all-websites-")
    assert name.endswith(".xlsx")
    wb = load_workbook(BytesIO(payload))
    assert wb.sheetnames == ["Overview", "Problems to fix"]
    assert wb["Overview"]["A4"].value == "Example"
    assert wb["Overview"]["F3"].value == "SSL expires"
    assert wb["Overview"]["H3"].value == "Domain expires"
    assert wb["Overview"]["F4"].value == "2026-09-01 (30 days left)"
    assert wb["Overview"]["H4"].value == "2027-01-15 (166 days left)"
    assert wb["Overview"]["C5"].value == "Not checked yet"
    assert wb["Overview"]["F5"].value == "—"
    assert wb["Overview"]["H5"].value == "—"
    assert wb["Problems to fix"]["A2"].value == "Example"
    assert wb["Problems to fix"]["E2"].value == "Certificate expiring soon"

    monkeypatch.setattr(
        "whm.infrastructure.reports.downloads_folder",
        lambda: tmp_path,
    )
    saved = save_portfolio_report_to_downloads(
        [(site, result)], format="excel"
    )
    assert saved.exists()
    assert saved.parent == tmp_path


def test_save_report_to_downloads(tmp_path: Path, monkeypatch):
    site, result = _sample()
    monkeypatch.setattr(
        "whm.infrastructure.reports.downloads_folder",
        lambda: tmp_path,
    )
    saved = save_report_to_downloads(site, result, format="excel")
    assert saved.parent == tmp_path
    assert saved.suffix == ".xlsx"
    assert saved.exists()
    csv_saved = save_report_to_downloads(site, result, format="csv")
    assert csv_saved.suffix == ".csv"
    assert csv_saved.exists()
