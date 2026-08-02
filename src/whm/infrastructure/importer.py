"""Import website lists from Excel (.xlsx) or CSV — flexible sheet layouts."""

from __future__ import annotations

import csv
import io
import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable, Optional
from urllib.parse import urlparse

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

# Headers are normalized: "Client Website" -> "clientwebsite", "URL" -> "url".
URL_KEYS = (
    "url",
    "websiteurl",
    "websiteaddress",
    "siteurl",
    "webaddress",
    "webpage",
    "homepage",
    "domain",
    "hostname",
    "fqdn",
    "host",
    "link",
    "address",
    "www",
    "web",
    "primaryurl",
    "liveurl",
    "productionurl",
    "site",
    "website",
)
NAME_KEYS = (
    "display_name",
    "displayname",
    "websitename",
    "sitename",
    "name",
    "title",
    "label",
    "site_name",
    "brand",
)
CUSTOMER_KEYS = (
    "customer",
    "customername",
    "client",
    "clientwebsite",  # spreadsheet header: "Client Website"
    "clientname",
    "company",
    "account",
    "organisation",
    "organization",
    "business",
)
DKIM_KEYS = ("dkim_selectors", "dkim", "selectors", "dkimselector")
INTERVAL_KEYS = ("check_interval", "interval", "schedule", "frequency")

_HEADER_HINTS = frozenset(URL_KEYS) | frozenset(NAME_KEYS) | frozenset(CUSTOMER_KEYS)
_SKIP_TEXT_PREFIXES = (
    "fill in",
    "only ",
    "note:",
    "notes:",
    "instruction",
    "example:",
    "tip:",
    "leave other",
    "use import",
)


@dataclass
class ImportRow:
    url: str
    display_name: str = ""
    customer: str = ""
    dkim_selectors: str = "s1,s2,em,default"
    check_interval: str = "manual"
    source_line: int = 0


IMPORT_TIP = (
    "Each row needs a real web address like example.com or https://example.com. "
    "Company names alone (for example “Amazon”) are not enough — put the domain "
    "in another column."
)


@dataclass
class ImportResult:
    added: list[str] = field(default_factory=list)
    skipped: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    @property
    def title(self) -> str:
        if self.added and not self.errors:
            return "Import complete"
        if self.added and self.errors:
            return "Import partly complete"
        if self.skipped and not self.added and not self.errors:
            return "Nothing new to add"
        if self.errors:
            return "Couldn’t import websites"
        return "No websites imported"

    @property
    def tone(self) -> str:
        """ok | warn | error — for UI toasts."""
        if self.added and not self.errors:
            return "ok"
        if self.added and self.errors:
            return "warn"
        if self.skipped and not self.added and not self.errors:
            return "ok"
        return "error"

    @property
    def tip(self) -> str:
        if self.errors or (not self.added and not self.skipped):
            return IMPORT_TIP
        if self.skipped and self.added:
            return "Skipped rows were already on your list."
        return ""

    @property
    def summary(self) -> str:
        n_add = len(self.added)
        n_skip = len(self.skipped)
        n_err = len(self.errors)
        site = "website" if n_add == 1 else "websites"

        if n_add and not n_err and not n_skip:
            return f"Imported {n_add} {site}."
        if n_add and not n_err and n_skip:
            return (
                f"Imported {n_add} {site}. "
                f"{n_skip} already on your list — left as is."
            )
        if n_add and n_err:
            problem = "problem" if n_err == 1 else "problems"
            return (
                f"Imported {n_add} {site}, but skipped {n_err} {problem}. "
                f"{self.errors[0]}"
            )
        if n_skip and not n_add and not n_err:
            return (
                f"All {n_skip} in this file are already on your list — nothing new to add."
            )
        if n_err == 1 and not n_add:
            return self.errors[0]
        if n_err and not n_add:
            return (
                f"Couldn’t import any websites ({n_err} problems). {self.errors[0]}"
            )
        return "No websites imported from that file."

    def as_api_dict(self) -> dict:
        return {
            "title": self.title,
            "summary": self.summary,
            "tip": self.tip,
            "tone": self.tone,
            "added": self.added,
            "skipped": self.skipped,
            "errors": self.errors,
            "added_count": len(self.added),
            "skipped_count": len(self.skipped),
            "error_count": len(self.errors),
        }


def friendly_parse_error(exc: BaseException, filename: str = "") -> str:
    """Turn technical import/parse failures into plain language."""
    name = (filename or "that file").strip() or "that file"
    text = str(exc) or exc.__class__.__name__
    lower = text.lower()
    if name.lower().endswith(".xls") or "old .xls" in lower:
        return (
            f"“{name}” is an older Excel format (.xls). "
            "Open it in Excel, choose File → Save As → Excel Workbook (.xlsx) "
            "or CSV, then import again."
        )
    if "no worksheet" in lower:
        return (
            f"“{name}” doesn’t look like it has a spreadsheet sheet. "
            "Try saving as .xlsx or CSV and import again."
        )
    if "zip" in lower or "not a valid" in lower and "xlsx" in lower:
        return (
            f"Couldn’t open “{name}” as an Excel workbook. "
            "Save it as .xlsx or CSV from Excel, then try again."
        )
    if "permission" in lower:
        return (
            f"Couldn’t read “{name}” — it may be open in Excel. "
            "Close the file and try again."
        )
    if "codec" in lower or "decode" in lower or "utf" in lower:
        return (
            f"Couldn’t read the text in “{name}”. "
            "In Excel, use Save As → CSV UTF-8, then import that file."
        )
    # Keep short custom ValueErrors; wrap the rest.
    if isinstance(exc, ValueError) and len(text) < 220:
        return text
    return (
        f"Couldn’t import “{name}”. "
        "Save as .xlsx or CSV and try again. "
        f"({text[:120]})"
    )


def _norm_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").strip().lower())


def _pick(row: dict[str, str], keys: Iterable[str]) -> str:
    for key in keys:
        if key in row and row[key].strip():
            return row[key].strip()
    return ""


def _looks_like_website(value: str) -> bool:
    """True if value could be a domain/URL — rejects brand names and notes."""
    text = (value or "").strip().strip("'\"")
    if not text:
        return False
    lower = text.lower()
    if any(lower.startswith(p) for p in _SKIP_TEXT_PREFIXES):
        return False
    if len(text) > 253:
        return False
    # Brand labels like "Amazon" / "Apple" — no dot, not a URL.
    if " " in text and "://" not in text:
        return False
    candidate = text
    if "://" not in candidate:
        candidate = "https://" + candidate
    host = (urlparse(candidate).hostname or "").strip().lower().rstrip(".")
    if not host or "." not in host:
        return False
    # Need a real-looking TLD (amazon.com, mybiz.co.za).
    labels = [p for p in host.split(".") if p]
    if len(labels) < 2:
        return False
    tld = labels[-1]
    if not re.fullmatch(r"[a-z]{2,24}", tld):
        return False
    if _norm_header(text) in _HEADER_HINTS:
        return False
    return True


def _best_url_in_cells(cells: Iterable[str]) -> str:
    candidates = [c.strip() for c in cells if _looks_like_website(c)]
    if not candidates:
        return ""
    # Prefer explicit URLs, then longer hostnames.
    candidates.sort(key=lambda c: (0 if "://" in c else 1, -len(c)))
    return candidates[0]


def _find_header_row(matrix: list[list[str]]) -> Optional[int]:
    """Locate the header row by column labels — not by fixed position."""
    scan_limit = min(len(matrix), 40)
    best: Optional[int] = None
    best_score = 0
    for idx in range(scan_limit):
        norms = [_norm_header(c) for c in matrix[idx]]
        score = 0
        if any(h in URL_KEYS for h in norms):
            score += 3
        if any(h in NAME_KEYS or h in CUSTOMER_KEYS for h in norms):
            score += 2
        if any(h in DKIM_KEYS or h in INTERVAL_KEYS for h in norms):
            score += 1
        joined = " ".join(norms)
        if "http" in joined or any(
            "." in (c or "") and " " not in (c or "") for c in matrix[idx]
        ):
            if score < 3:
                continue
        if score > best_score:
            best_score = score
            best = idx
    if best is not None and best_score >= 3:
        return best
    for idx in range(scan_limit):
        norms = [_norm_header(c) for c in matrix[idx]]
        if any(h in URL_KEYS for h in norms):
            return idx
    return None


def _pad_row(row: list[str], width: int) -> list[str]:
    if len(row) >= width:
        return row[:width]
    return row + [""] * (width - len(row))


def _column_website_scores(matrix: list[list[str]], start: int, width: int) -> list[int]:
    scores = [0] * width
    for i in range(start, len(matrix)):
        line = _pad_row(matrix[i], width)
        for j in range(width):
            if _looks_like_website(line[j]):
                scores[j] += 1
    return scores


def _header_url_bonus(headers: list[str], width: int) -> list[int]:
    bonus = [0] * width
    for j in range(min(width, len(headers))):
        h = headers[j]
        if h in ("url", "websiteurl", "siteurl", "domain", "hostname", "link"):
            bonus[j] += 5
        elif h in URL_KEYS:
            bonus[j] += 2
    return bonus


def _pick_url_column(
    scores: list[int], header_bonus: list[int]
) -> Optional[int]:
    if not scores:
        return None
    ranked = sorted(
        range(len(scores)),
        key=lambda j: (scores[j] + header_bonus[j], scores[j], header_bonus[j]),
        reverse=True,
    )
    best = ranked[0]
    if scores[best] <= 0 and header_bonus[best] <= 0:
        return None
    # If header says "website" but almost no real domains there, and another
    # column is full of domains, prefer the data column.
    if scores[best] == 0:
        data_best = max(range(len(scores)), key=lambda j: scores[j])
        if scores[data_best] > 0:
            return data_best
    return best


def _text_label(value: str) -> str:
    text = (value or "").strip()
    if not text:
        return ""
    lower = text.lower()
    if any(lower.startswith(p) for p in _SKIP_TEXT_PREFIXES):
        return ""
    if _looks_like_website(text):
        return ""
    if _norm_header(text) in _HEADER_HINTS:
        return ""
    return text


def _rows_from_matrix_with_lines(
    numbered: list[tuple[int, list[str]]],
) -> list[ImportRow]:
    matrix = [row for _, row in numbered]
    line_nos = [n for n, _ in numbered]
    if not matrix:
        return []

    width = max((len(r) for r in matrix), default=0)
    if width == 0:
        return []

    header_idx = _find_header_row(matrix)
    data_start = (header_idx + 1) if header_idx is not None else 0
    headers = (
        [_norm_header(h) for h in _pad_row(matrix[header_idx], width)]
        if header_idx is not None
        else [""] * width
    )

    scores = _column_website_scores(matrix, data_start, width)
    url_col = _pick_url_column(scores, _header_url_bonus(headers, width))

    name_cols = [j for j, h in enumerate(headers) if h in NAME_KEYS]
    customer_cols = [j for j, h in enumerate(headers) if h in CUSTOMER_KEYS]
    dkim_cols = [j for j, h in enumerate(headers) if h in DKIM_KEYS]
    interval_cols = [j for j, h in enumerate(headers) if h in INTERVAL_KEYS]

    rows: list[ImportRow] = []
    for i in range(data_start, len(matrix)):
        line = _pad_row(matrix[i], width)
        if not any(line):
            continue

        url = ""
        if url_col is not None:
            url = line[url_col].strip()
        if not _looks_like_website(url):
            url = _best_url_in_cells(line)
        if not url:
            # Row has only labels / notes — skip quietly (not an error).
            continue

        customer = ""
        for j in customer_cols:
            customer = _text_label(line[j])
            if customer:
                break
        display_name = ""
        for j in name_cols:
            display_name = _text_label(line[j])
            if display_name:
                break
        if not display_name:
            display_name = customer
        # If still empty, use a non-URL text cell as a friendly name.
        if not display_name:
            for cell in line:
                label = _text_label(cell)
                if label:
                    display_name = label
                    break

        dkim = "s1,s2,em,default"
        for j in dkim_cols:
            if line[j].strip():
                dkim = line[j].strip()
                break
        interval = "manual"
        for j in interval_cols:
            if line[j].strip():
                interval = line[j].strip()
                break

        rows.append(
            ImportRow(
                url=url,
                display_name=display_name,
                customer=customer,
                dkim_selectors=dkim,
                check_interval=interval,
                source_line=line_nos[i],
            )
        )
    return rows


def parse_csv_bytes(data: bytes) -> list[ImportRow]:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    numbered: list[tuple[int, list[str]]] = []
    for idx, row in enumerate(reader, start=1):
        cells = [cell.strip() for cell in row]
        if any(cells):
            numbered.append((idx, cells))
    return _rows_from_matrix_with_lines(numbered)


def _col_to_index(cell_ref: str) -> int:
    letters = re.match(r"[A-Z]+", cell_ref.upper())
    if not letters:
        return 0
    value = 0
    for ch in letters.group(0):
        value = value * 26 + (ord(ch) - 64)
    return value - 1


def _parse_xlsx_openpyxl(data: bytes) -> list[ImportRow]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    combined: list[ImportRow] = []
    try:
        for sheet in wb.worksheets:
            numbered: list[tuple[int, list[str]]] = []
            for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cells = [
                    "" if c is None else str(c).strip()
                    for c in row
                ]
                # Trim trailing empties but keep interior columns.
                while cells and not cells[-1]:
                    cells.pop()
                if any(cells):
                    numbered.append((idx, cells))
            combined.extend(_rows_from_matrix_with_lines(numbered))
    finally:
        wb.close()
    return combined


def _parse_xlsx_minimal(data: bytes) -> list[ImportRow]:
    """Minimal XLSX reader (first sheet) — no openpyxl required."""
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for si in root.findall("m:si", NS):
                texts = [t.text or "" for t in si.findall(".//m:t", NS)]
                shared.append("".join(texts))

        sheet_names = sorted(
            n for n in archive.namelist() if re.match(r"xl/worksheets/sheet\d+\.xml", n)
        )
        if not sheet_names:
            raise ValueError("No worksheet found in Excel file")

        combined: list[ImportRow] = []
        for sheet_name in sheet_names:
            root = ET.fromstring(archive.read(sheet_name))
            numbered: list[tuple[int, list[str]]] = []
            for row in root.findall("m:sheetData/m:row", NS):
                row_num = int(row.attrib.get("r", len(numbered) + 1))
                cells: dict[int, str] = {}
                for cell in row.findall("m:c", NS):
                    ref = cell.attrib.get("r", "A1")
                    idx = _col_to_index(ref)
                    cell_type = cell.attrib.get("t")
                    if cell_type == "inlineStr":
                        texts = [t.text or "" for t in cell.findall(".//m:t", NS)]
                        raw = "".join(texts)
                    else:
                        value_el = cell.find("m:v", NS)
                        if value_el is None or value_el.text is None:
                            raw = ""
                        elif cell_type == "s":
                            si = int(value_el.text)
                            raw = shared[si] if 0 <= si < len(shared) else ""
                        else:
                            raw = value_el.text
                    cells[idx] = raw
                if not cells:
                    continue
                width = max(cells) + 1
                line = [str(cells.get(i, "")).strip() for i in range(width)]
                if any(line):
                    numbered.append((row_num, line))
            combined.extend(_rows_from_matrix_with_lines(numbered))
        return combined


def parse_xlsx_bytes(data: bytes) -> list[ImportRow]:
    try:
        rows = _parse_xlsx_openpyxl(data)
        if rows:
            return rows
    except Exception:
        pass
    return _parse_xlsx_minimal(data)


def parse_import_file(filename: str, data: bytes) -> list[ImportRow]:
    name = filename.lower().strip()
    try:
        if name.endswith(".csv") or name.endswith(".txt"):
            return parse_csv_bytes(data)
        if name.endswith(".xlsx"):
            return parse_xlsx_bytes(data)
        if name.endswith(".xls"):
            raise ValueError(
                "Old .xls format is not supported. In Excel: File → Save As → "
                "Excel Workbook (.xlsx) or CSV, then import again."
            )
        if data[:2] == b"PK":
            return parse_xlsx_bytes(data)
        return parse_csv_bytes(data)
    except ValueError:
        raise
    except Exception as exc:  # noqa: BLE001
        raise ValueError(friendly_parse_error(exc, filename)) from exc


def parse_import_path(path: str | Path) -> list[ImportRow]:
    file_path = Path(path)
    return parse_import_file(file_path.name, file_path.read_bytes())


def apply_import(
    rows: list[ImportRow],
    *,
    existing_domains: set[str],
    add_customer,
    add_website,
    extract_domain,
) -> ImportResult:
    """
    Import rows using injected service callables to keep this layer free of UI.

    add_customer(name) -> object with .id
    add_website(...) -> website with .domain
    extract_domain(url) -> str
    """
    result = ImportResult()
    seen = set(existing_domains)

    if not rows:
        result.errors.append(
            "No website addresses found in that file. "
            "Add a column with domains or links (for example amazon.com or "
            "https://example.com), then import again."
        )
        return result

    for row in rows:
        if not (row.url or "").strip():
            continue
        if not _looks_like_website(row.url):
            snippet = row.url.replace("\n", " ").strip()
            if len(snippet) > 40:
                snippet = snippet[:37] + "…"
            result.errors.append(
                f"Row {row.source_line}: “{snippet}” doesn’t look like a web address"
            )
            continue
        try:
            domain = extract_domain(row.url)
        except Exception as exc:  # noqa: BLE001
            result.errors.append(
                f"Row {row.source_line}: couldn’t use that address ({exc})"
            )
            continue
        if domain in seen:
            result.skipped.append(domain)
            continue
        customer_id = None
        customer_name = (row.customer or "").strip()
        if customer_name and any(
            customer_name.lower().startswith(p) for p in _SKIP_TEXT_PREFIXES
        ):
            customer_name = ""
        if customer_name:
            customer_id = add_customer(customer_name).id
        try:
            site = add_website(
                url=row.url,
                display_name=row.display_name,
                customer_id=customer_id,
                dkim_selectors=row.dkim_selectors or "s1,s2,em,default",
                check_interval=row.check_interval or "manual",
            )
        except Exception as exc:  # noqa: BLE001
            result.errors.append(
                f"Row {row.source_line}: couldn’t save {row.url} ({exc})"
            )
            continue
        seen.add(site.domain)
        result.added.append(site.domain)
    return result
