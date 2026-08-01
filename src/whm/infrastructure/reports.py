"""Export clear Excel or CSV health reports for support staff / customers."""

from __future__ import annotations

import csv
import io
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from whm.domain.models import FindingStatus, HealthCheckResult, Website
from whm.presentation.copy import (
    category_plain,
    finding_plain,
    overall_summary,
    risk_plain,
    status_plain,
)

_IGNORED_REPORT_CATEGORIES = frozenset(
    {
        "security",
        "performance",
        "smtp",
        "spf",
        "dkim",
        "dmarc",
        "mx",
        "sendgrid",
    }
)
_ACTION_STATUSES = frozenset(
    {
        FindingStatus.MISSING,
        FindingStatus.INCORRECT,
        FindingStatus.INCONCLUSIVE,
    }
)

_HEADER_FILL = PatternFill("solid", fgColor="1B2430")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1B2430")
_LABEL_FONT = Font(name="Calibri", bold=True, size=11, color="334155")
_BODY_FONT = Font(name="Calibri", size=11, color="0F172A")
_GOOD_FILL = PatternFill("solid", fgColor="D1FAE5")
_WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
_BAD_FILL = PatternFill("solid", fgColor="FEE2E2")
_MUTED_FILL = PatternFill("solid", fgColor="E2E8F0")
_THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _safe_name(domain: str) -> str:
    return "".join(ch if ch.isalnum() or ch in "-._" else "_" for ch in domain)


def _report_base_name(website: Website) -> str:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"{_safe_name(website.domain)}-{stamp}"


def _report_findings(result: HealthCheckResult):
    return [f for f in result.findings if f.category not in _IGNORED_REPORT_CATEGORIES]


def _action_findings(result: HealthCheckResult):
    return [f for f in _report_findings(result) if f.status in _ACTION_STATUSES]


def _status_fill(label: str) -> PatternFill:
    lower = label.lower()
    if "looks good" in lower or lower == "ok" or lower == "low":
        return _GOOD_FILL
    if (
        "worth a look" in lower
        or "attention" in lower
        or "needs fixing" in lower
        or "review" in lower
        or "medium" in lower
    ):
        return _WARN_FILL
    if (
        "needs a fix" in lower
        or "wrong" in lower
        or "missing" in lower
        or "not set up" in lower
        or "high" in lower
    ):
        return _BAD_FILL
    return _MUTED_FILL


def _set_widths(ws, widths: list[float]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _style_header_row(ws, row: int, columns: int) -> None:
    for col in range(1, columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = _THIN


def _write_kv(ws, row: int, label: str, value: str, *, colour_value: bool = False) -> int:
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = _LABEL_FONT
    label_cell.alignment = _WRAP
    label_cell.border = _THIN
    value_cell = ws.cell(row=row, column=2, value=value)
    value_cell.font = _BODY_FONT
    value_cell.alignment = _WRAP
    value_cell.border = _THIN
    if colour_value:
        value_cell.fill = _status_fill(value)
    return row + 1


def render_excel(website: Website, result: HealthCheckResult) -> bytes:
    """Build a two-sheet .xlsx: Summary + Problems to fix."""
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Website Health Manager — Report"
    summary["A1"].font = _TITLE_FONT
    summary.merge_cells("A1:B1")

    row = 3
    row = _write_kv(summary, row, "Website", website.display_name)
    row = _write_kv(summary, row, "Domain", website.domain)
    row = _write_kv(summary, row, "URL", website.url)
    row = _write_kv(
        summary,
        row,
        "Checked at",
        result.checked_at.strftime("%Y-%m-%d %H:%M UTC"),
    )
    row = _write_kv(
        summary,
        row,
        "Overall",
        status_plain(result.overall_status),
        colour_value=True,
    )
    row = _write_kv(
        summary,
        row,
        "Risk",
        risk_plain(result.risk_level),
        colour_value=True,
    )
    row += 1
    summary.cell(row=row, column=1, value="Area scores").font = _TITLE_FONT
    row += 1
    for label, status in (
        ("Website opens", result.website_status),
        ("Security certificate (SSL)", result.ssl_status),
        ("Domain registration", result.domain_status),
        ("Web address settings (DNS)", result.dns_status),
    ):
        row = _write_kv(
            summary,
            row,
            label,
            status_plain(status),
            colour_value=True,
        )

    row += 1
    row = _write_kv(
        summary,
        row,
        "In plain words",
        overall_summary(result.overall_status, website.display_name),
    )
    actions = _action_findings(result)
    row = _write_kv(
        summary,
        row,
        "Items to fix",
        str(len(actions)),
    )
    _set_widths(summary, [34, 72])
    summary.row_dimensions[1].height = 22
    for r in range(3, row):
        summary.row_dimensions[r].height = 28
    summary.freeze_panes = "A3"

    problems = wb.create_sheet("Problems to fix")
    headers = ["Area", "Status", "Problem", "Details", "What to do"]
    for col, header in enumerate(headers, start=1):
        problems.cell(row=1, column=col, value=header)
    _style_header_row(problems, 1, len(headers))

    if not actions:
        problems.cell(
            row=2,
            column=1,
            value="Nothing to fix — everything we checked looks fine.",
        )
        problems.merge_cells("A2:E2")
        problems["A2"].font = _BODY_FONT
        problems["A2"].fill = _GOOD_FILL
        problems["A2"].alignment = _WRAP
    else:
        for idx, finding in enumerate(actions, start=2):
            values = [
                category_plain(finding.category),
                finding_plain(finding.status),
                finding.title,
                finding.message,
                finding.recommendation or "Ask the hosting provider to fix this.",
            ]
            for col, value in enumerate(values, start=1):
                cell = problems.cell(row=idx, column=col, value=value)
                cell.font = _BODY_FONT
                cell.alignment = _WRAP
                cell.border = _THIN
                if col == 2:
                    cell.fill = _status_fill(str(value))
            problems.row_dimensions[idx].height = max(36, 18 * (1 + len(values[3]) // 60))

    _set_widths(problems, [28, 14, 28, 44, 44])
    problems.auto_filter.ref = f"A1:E{max(2, len(actions) + 1)}"
    problems.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def render_csv(website: Website, result: HealthCheckResult) -> str:
    """Flat, readable CSV: summary rows, then problems to fix."""
    handle = io.StringIO()
    writer = csv.writer(handle)
    writer.writerow(["Section", "Field", "Value"])
    writer.writerow(["Summary", "Website", website.display_name])
    writer.writerow(["Summary", "Domain", website.domain])
    writer.writerow(["Summary", "URL", website.url])
    writer.writerow(
        ["Summary", "Checked at", result.checked_at.strftime("%Y-%m-%d %H:%M UTC")]
    )
    writer.writerow(["Summary", "Overall", status_plain(result.overall_status)])
    writer.writerow(["Summary", "Risk", risk_plain(result.risk_level)])
    writer.writerow(
        ["Summary", "Website opens", status_plain(result.website_status)]
    )
    writer.writerow(
        [
            "Summary",
            "Security certificate (SSL)",
            status_plain(result.ssl_status),
        ]
    )
    writer.writerow(
        ["Summary", "Domain registration", status_plain(result.domain_status)]
    )
    writer.writerow(
        [
            "Summary",
            "Web address settings (DNS)",
            status_plain(result.dns_status),
        ]
    )
    writer.writerow(
        [
            "Summary",
            "In plain words",
            overall_summary(result.overall_status, website.display_name),
        ]
    )

    writer.writerow([])
    writer.writerow(["Area", "Status", "Problem", "Details", "What to do"])
    actions = _action_findings(result)
    if not actions:
        writer.writerow(
            [
                "",
                "",
                "Nothing to fix",
                "Everything we checked looks fine.",
                "",
            ]
        )
    else:
        for finding in actions:
            writer.writerow(
                [
                    category_plain(finding.category),
                    finding_plain(finding.status),
                    finding.title,
                    finding.message,
                    finding.recommendation
                    or "Ask the hosting provider to fix this.",
                ]
            )
    return handle.getvalue()


def export_excel(path: Path, website: Website, result: HealthCheckResult) -> Path:
    path.write_bytes(render_excel(website, result))
    return path


def export_csv(path: Path, website: Website, result: HealthCheckResult) -> Path:
    path.write_text(render_csv(website, result), encoding="utf-8-sig")
    return path


def build_excel_report(website: Website, result: HealthCheckResult) -> tuple[str, bytes]:
    base = _report_base_name(website)
    return f"{base}-report.xlsx", render_excel(website, result)


def build_csv_report(website: Website, result: HealthCheckResult) -> tuple[str, bytes]:
    base = _report_base_name(website)
    return f"{base}-report.csv", render_csv(website, result).encode("utf-8-sig")


def downloads_folder() -> Path:
    """User Downloads folder (Windows-friendly), creating it if needed."""
    home = Path.home()
    candidates = [
        home / "Downloads",
        Path.home() / "OneDrive" / "Downloads",
    ]
    for path in candidates:
        if path.is_dir():
            return path
    target = home / "Downloads"
    target.mkdir(parents=True, exist_ok=True)
    return target


def save_report_to_downloads(
    website: Website,
    result: HealthCheckResult,
    *,
    format: str = "excel",
) -> Path:
    """Write Excel or CSV straight into the user's Downloads folder."""
    fmt = (format or "excel").strip().lower()
    if fmt == "csv":
        filename, payload = build_csv_report(website, result)
    else:
        filename, payload = build_excel_report(website, result)
    path = downloads_folder() / filename
    path.write_bytes(payload)
    return path


def render_portfolio_excel(
    rows: list[tuple[Website, HealthCheckResult | None]],
) -> bytes:
    """All-websites workbook: Overview + Problems across the portfolio."""
    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"
    overview["A1"] = "Website Health Manager — All websites"
    overview["A1"].font = _TITLE_FONT
    overview.merge_cells("A1:I1")

    headers = [
        "Website",
        "Domain",
        "Overall",
        "Web",
        "SSL",
        "Domain reg.",
        "DNS",
        "Last checked",
        "Items to fix",
    ]
    for col, header in enumerate(headers, start=1):
        overview.cell(row=3, column=col, value=header)
    _style_header_row(overview, 3, len(headers))

    problems = wb.create_sheet("Problems to fix")
    problem_headers = ["Website", "Domain", "Area", "Status", "Problem", "Details", "What to do"]
    for col, header in enumerate(problem_headers, start=1):
        problems.cell(row=1, column=col, value=header)
    _style_header_row(problems, 1, len(problem_headers))

    problem_row = 2
    total_actions = 0
    for idx, (site, result) in enumerate(rows, start=4):
        if result is None:
            values = [
                site.display_name,
                site.domain,
                "Not checked yet",
                "—",
                "—",
                "—",
                "—",
                "Never",
                "0",
            ]
        else:
            actions = _action_findings(result)
            total_actions += len(actions)
            values = [
                site.display_name,
                site.domain,
                status_plain(result.overall_status),
                status_plain(result.website_status),
                status_plain(result.ssl_status),
                status_plain(result.domain_status),
                status_plain(result.dns_status),
                result.checked_at.strftime("%Y-%m-%d %H:%M"),
                str(len(actions)),
            ]
            for finding in actions:
                prow = [
                    site.display_name,
                    site.domain,
                    category_plain(finding.category),
                    finding_plain(finding.status),
                    finding.title,
                    finding.message,
                    finding.recommendation
                    or "Ask the hosting provider to fix this.",
                ]
                for col, value in enumerate(prow, start=1):
                    cell = problems.cell(row=problem_row, column=col, value=value)
                    cell.font = _BODY_FONT
                    cell.alignment = _WRAP
                    cell.border = _THIN
                    if col == 4:
                        cell.fill = _status_fill(str(value))
                problem_row += 1

        for col, value in enumerate(values, start=1):
            cell = overview.cell(row=idx, column=col, value=value)
            cell.font = _BODY_FONT
            cell.alignment = _WRAP
            cell.border = _THIN
            if col in {3, 4, 5, 6, 7}:
                cell.fill = _status_fill(str(value))

    if problem_row == 2:
        problems.cell(
            row=2,
            column=1,
            value="Nothing to fix across the checked websites.",
        )
        problems.merge_cells("A2:G2")
        problems["A2"].font = _BODY_FONT
        problems["A2"].fill = _GOOD_FILL

    overview.cell(row=2, column=1, value=f"Sites: {len(rows)} · Items to fix: {total_actions}")
    overview["A2"].font = _BODY_FONT
    _set_widths(overview, [22, 22, 14, 12, 12, 14, 12, 16, 12])
    _set_widths(problems, [22, 22, 22, 12, 26, 40, 40])
    overview.freeze_panes = "A4"
    problems.freeze_panes = "A2"
    problems.auto_filter.ref = f"A1:G{max(2, problem_row - 1)}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def render_portfolio_csv(
    rows: list[tuple[Website, HealthCheckResult | None]],
) -> str:
    handle = io.StringIO()
    writer = csv.writer(handle)
    writer.writerow(
        [
            "Website",
            "Domain",
            "Overall",
            "Web",
            "SSL",
            "Domain reg.",
            "DNS",
            "Last checked",
            "Items to fix",
        ]
    )
    for site, result in rows:
        if result is None:
            writer.writerow(
                [
                    site.display_name,
                    site.domain,
                    "Not checked yet",
                    "—",
                    "—",
                    "—",
                    "—",
                    "Never",
                    "0",
                ]
            )
        else:
            actions = _action_findings(result)
            writer.writerow(
                [
                    site.display_name,
                    site.domain,
                    status_plain(result.overall_status),
                    status_plain(result.website_status),
                    status_plain(result.ssl_status),
                    status_plain(result.domain_status),
                    status_plain(result.dns_status),
                    result.checked_at.strftime("%Y-%m-%d %H:%M"),
                    str(len(actions)),
                ]
            )

    writer.writerow([])
    writer.writerow(
        ["Website", "Domain", "Area", "Status", "Problem", "Details", "What to do"]
    )
    any_problems = False
    for site, result in rows:
        if result is None:
            continue
        for finding in _action_findings(result):
            any_problems = True
            writer.writerow(
                [
                    site.display_name,
                    site.domain,
                    category_plain(finding.category),
                    finding_plain(finding.status),
                    finding.title,
                    finding.message,
                    finding.recommendation
                    or "Ask the hosting provider to fix this.",
                ]
            )
    if not any_problems:
        writer.writerow(
            ["", "", "", "", "Nothing to fix", "Everything checked looks fine.", ""]
        )
    return handle.getvalue()


def build_portfolio_excel_report(
    rows: list[tuple[Website, HealthCheckResult | None]],
) -> tuple[str, bytes]:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"whm-all-websites-{stamp}.xlsx", render_portfolio_excel(rows)


def build_portfolio_csv_report(
    rows: list[tuple[Website, HealthCheckResult | None]],
) -> tuple[str, bytes]:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"whm-all-websites-{stamp}.csv", render_portfolio_csv(rows).encode("utf-8-sig")


def save_portfolio_report_to_downloads(
    rows: list[tuple[Website, HealthCheckResult | None]],
    *,
    format: str = "excel",
) -> Path:
    fmt = (format or "excel").strip().lower()
    if fmt == "csv":
        filename, payload = build_portfolio_csv_report(rows)
    else:
        filename, payload = build_portfolio_excel_report(rows)
    path = downloads_folder() / filename
    path.write_bytes(payload)
    return path


def default_export_paths(folder: Path, website: Website) -> dict[str, Path]:
    folder.mkdir(parents=True, exist_ok=True)
    base = _report_base_name(website)
    return {
        "excel": folder / f"{base}-report.xlsx",
        "csv": folder / f"{base}-report.csv",
    }
